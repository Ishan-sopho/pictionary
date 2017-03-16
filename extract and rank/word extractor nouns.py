import openpyxl
wb = openpyxl.load_workbook('Pictionary-Words.xlsx')
sheet = wb.get_sheet_by_name('Table 1')
noun,easy,medium,hard=[],[],[],[]
feasy=open(r'easynouns.txt','w')
fmedium=open(r'mediumnouns.txt','w')
fhard=open(r'hardnouns.txt','w')
for i in range (1,4,1):
    for j in range (3,23,1):
        feasy.write((sheet.cell(row=j,column=i).value).encode('utf-8'))
        feasy.write('\n')
    for k in range (24,67,1):
        fmedium.write((sheet.cell(row=k, column=i).value).encode('utf-8'))
        fmedium.write('\n')
    for l in range (68,111,1):
        fhard.write((sheet.cell(row=l, column=i).value).encode('utf-8'))
        fhard.write('\n')
feasy.close()
fmedium.close()
fhard.close()
