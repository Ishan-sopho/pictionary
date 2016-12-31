import openpyxl
wb = openpyxl.load_workbook('Pictionary-Words.xlsx')
sheet = wb.get_sheet_by_name('Table 1')
noun,easy,medium,hard=[],[],[],[]
for i in range (1,4,1):
    for j in range (3,23,1):
        easy.append(sheet.cell(row=j,column=i).value)
    noun.append(easy)
    for k in range (24,67,1):
        medium.append(sheet.cell(row=k,column=i).value)
    noun.append(medium)
    for l in range (68,111,1):
        hard.append(sheet.cell(row=l,column=i).value)
    noun.append(hard)
for a in noun:
    for b in a:
        print b
import random
word=0
while (raw_input(':')!='stop'):
    n=random.randint(1,10)
    if n<=3:
        word=random.randint(0,59)
        print noun[0][word]
    elif n<=7:
        word=random.randint(0,128)
        print noun[1][word]
    else:
        word=random.randint(0,128)
        print noun[2][word]
